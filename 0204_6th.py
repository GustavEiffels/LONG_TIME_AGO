# Threading
# import threading
# import time
#
# print("--------- 공유자원 수정 문제 발생 ")
#
# # Sleep 을 기재하는 이유
# # 수업중에 하는 것들은 Delay 가 발생하기 어렵기 때문에 추가하는 것
# # 실제로 Thread  사용시 Sleep 을 사용하지 않을 수 있다.
#
# # 공유 자원을 생성
# g_count = 0
#
# # 상호배제를 위한 인스턴스 생성
# lock = threading.Lock()
#
# print("Thread Class 상속 ---- Thread Class 생성 ")
#
#
# # 파이토치에서 상속받아서 많이 사용
# # 자유도가 높다는 장점이 있다 .
# class ThreadEx(threading.Thread):
#     def run(self):
#         # 파이썬에서 함수 외부의 변수를 사용
#         global g_count
#         global lock
#
#         for i in range(0, 10, 1):
#             # 다른 Thread 가 작업할 수 없도록 한다
#             lock.acquire()
#             print('id={0} 증가하기 전 ---> {1}'.format(self.getName(), g_count))
#             g_count = g_count + 1
#             # Thread 전환을 위해서 잠시 대기
#             time.sleep(0.5)
#             print('id={0} 증가하기 후 ---> {1}'.format(self.getName(), g_count))
#             # 다른 Thread 가 작업할 수 있도록 한다
#             lock.release()
#             time.sleep(0.5)


#  Thread를 생성해서 실행

# for i in range(2):
#     th = ThreadEx()
#     th.start()

print("File Handling ------------------------------")
print("파일에 읽고 쓰기 ")
# 현재 디렉토리 확인
import os

print(os.getcwd())

try:
    # file 을 쓰기 모드로 개방
    # 'a' --> 추가
    # 'w' --> 기록만
    file = open('./test.txt', 'w')
    # file 에 기록
    file.write("Hello FileHandling")
    file.write("₩n")
    li = ["I", "wanna", "be", "a", "richMan",
          "and", "being", "a", "irreplaceable", "person", "할 수 있다."]
    # list 의 내용을   ₩n 을 중간에
    file.write("\n".join(li))

    file.close()
except Exception as e:
    print(e)

print("파일을 읽기 모드로 개방 -----------")
import time

try:
    # 파일을 읽기 모드로 개방
    file = open('./test.txt', 'r')

    # 전체 읽어오기 --> 용량이 크면 메모리 부족으로 읽어오지 못함
    # print(file.read())

    # 줄 단위 읽기
    for line in file:
        print(line)

        time.sleep(0.2)

    file.close()
except Exception as e:
    print(e)

# 예외처리하지 않고 닫기


print("CSV 파일 핸들링 -------------------")
# 파일 읽기를 이용해서 csv 파일 읽기 ---> 모든 데이터를 분할해서 list 로 생성

with open("./data/singer.csv", 'r') as f:
    # 전체 읽기
    content = f.read()

    # 구분자 단위로 분할
    data = content.split('.')

    # 자료형 확인
    print(type(data))

    # 전체 출력
    for imsi in data:
        print(data)

# CSV 모듈을 이용해서 읽기 csv.reader(file 객체)를 분할해서 list 로 생성
import csv

# 한글이 깨지면 Encoding 옵션에 인코딩 방식을 설정
with open("./data/singer.csv", 'r', encoding="utf-8") as f:
    # csv 로 읽기
    # 구분자가 , 가 아니면 delimiter 에 구분자를 설정해주면 된다.

    data = csv.reader(f)

    # 자료형 확인
    print(type(data))

    # 전체 출력
    for imsi in data:
        print(imsi)

print("기록하기 -------------------------")
import csv

# newLine 옵션을이용해서 빈줄이 생기지 않도록 설정
# 한글이 깨지면 Encoding option에 Encoding 방식을 서렂ㅇ
with open("./data/singer.csv", 'a', newline='') as f:
    # csv로 기록
    wr = csv.writer(f)
    wr.writerow([4, "karina", "2000-01-01", "aespa"])

print("Byte 단위로 파일을 읽고 쓰기------------------------")
with open('./data/test.bin', 'wb') as f:
    f.write("Hello 커피".encode())

with open('./data/test.bin', 'rb') as f:
    # print(f.read())
    print(f.read().decode())

print("Serializable(직렬화) ----------------- 객체 단위로 파일에 읽고 쓰는 것 ")


class Vo:


    def setNum(self, num):
        self.num = num

    def setName(self, name):
        self.name = name

    def getNum(self):
        return self.num

    def getName(self):
        return self.name

    def toString(self):
        return f"번호:{self.num} 이름:{self.name}"

# 직렬화 할 인스턴스 생성
dto1 =Vo()
dto1.setNum(1)
dto1.setName("Nani")

dto2 =Vo()
dto2.setNum(2)
dto2.setName("SInde_Eru")

li = [dto1, dto2]

import pickle
with open('./data/data.dat', 'wb')as f:
    pickle.dump(li, f)

import pickle
with open('./data/data.dat', 'rb')as f:
    result = pickle.load(f)
    for Vo in result:
        print(Vo.toString())

print("excel 파일 읽기 ------------------------------------------------")
import openpyxl

# 엑셀 파일 포인터
wb = None

try:
    #엑셀 파일에 포인터를 생성
    wb = openpyxl.load_workbook('./data/score2.xlsx')

    # 엑셀 파일 경로 확인
    print(wb)

    # sheet 가져오기
    # 현재 활성화된 sheet를 가져온다
    ws = wb.active


    print("Horizon--------------------------")
    for row in ws.rows:
        if row[0].row == 1:
            ws.cell(row=row[0].row, column=5).value="합계"
            continue
        print("합계 구하기")
        total = row[1].value + row[2].value + row[3].value
        ws.cell(row=row[0].row, column=5).value = total

    # 파일에 기록
    wb.save('./data/score2.xlsx')
except Exception as e:
    print(e)

finally:
    if wb != None:
        wb.close()

print("mysql 접속 ---------------------------------")
import pymysql

con = pymysql.connect(host='127.0.0.1', port=3306,
                      user='singsiuk', password='ssw0304',
                      db='python', charset='utf8')

print(con)

con.close()